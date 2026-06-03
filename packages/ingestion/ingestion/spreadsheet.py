"""Spreadsheet ingestion: CSV (stdlib) and XLSX (openpyxl, MIT). Lossless, cells as strings."""

from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import load_workbook

from .models import IngestResult, add_table, rows_to_table


def ingest_csv(path: str | Path) -> IngestResult:
    path = Path(path)
    with path.open(newline="", encoding="utf-8-sig") as handle:
        raw = list(csv.reader(handle))
    result = IngestResult()
    table = rows_to_table(raw, caption=path.name)
    if table is not None:
        add_table(result, table, asset_id=path.stem, source=path.name, locator={})
    return result


def ingest_xlsx(path: str | Path) -> IngestResult:
    path = Path(path)
    result = IngestResult()
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        for sheet in workbook.worksheets:
            table = rows_to_table(
                sheet.iter_rows(values_only=True),
                caption=f"{path.name} · {sheet.title}",
            )
            if table is not None:
                add_table(
                    result,
                    table,
                    asset_id=f"{path.stem}:{sheet.title}",
                    source=path.name,
                    locator={"sheet": sheet.title},
                )
    finally:
        workbook.close()
    return result
